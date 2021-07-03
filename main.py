import csv
import glob
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl import styles

pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 100)
pd.set_option('display.width', 500)
pd.set_option('display.max_colwidth', 70)


def read_excel_sheets(xlsx_path):
    """Read all sheets of an Excel workbook and return a list with selected sheets names
    Args:
        xlsx_path (str): path with excel file to import
    """
    sheet_list = []
    xlsx = pd.read_excel(xlsx_path, engine='openpyxl', sheet_name=None)
    for page in list(xlsx.keys()):
        if (page == "Frontpage") or (page == "TOC") or (page == "Amendment Log") or (page == "Piping Insulation"):
            pass
        else:
            sheet_list.append(page)
    return sheet_list


def add_special_column(xlsx_path, output_csv_path):
    """Convert selected DataFrame to csv file with pointed parameters
    Args:
        xlsx_path (str): path with excel file to import
        output_csv_path (str): path, where csv file be created
    """
    sheets = read_excel_sheets(xlsx_path)
    dfs = []
    for idx, df in enumerate(range(len(sheets))):
        a = sheets[df]
        df = pd.read_excel(xlsx_path, sheets[df], engine='openpyxl', skiprows=6, index_col=None).dropna(
            how='all').dropna(axis=1, how='all')
        df["Sheet name"] = "{} {}".format((idx+3), a)
        dfs.append(df)
    df_final = pd.concat(dfs, axis=0, join='outer')
    name = os.path.splitext(xlsx_path)[0]
    name2 = name.split('\\')[1]
    df_final.to_csv(os.path.join(output_csv_path, "{}.csv".format(name2)), mode='w', sep=';',
                    quoting=csv.QUOTE_NONE, quotechar="", escapechar="\\", index=False)


def both_data_csv_output(input_directory, transient_csv_folder):
    """Export csv from selected directory to pointed folder by using previous functions
    Args:
        input_directory (str): path with excel files to import
        transient_csv_folder (str): path, where csv files be created
    """

    data_file_list = glob.glob(input_directory + "\*")
    [add_special_column(file, transient_csv_folder) for file in data_file_list]

    return "Exporting process done"


def csv_to_excel(transient_csv_file, transient_xlsx_file):
    """Convert selected DataFrame (created from csv files) to excel file.
    Args:
        transient_csv_file (str): path with csv files to import
        transient_xlsx_file (str): path, where xlsx file be created
    """

    csv_files = glob.glob(transient_csv_file + '/*.csv')
    for index, item in enumerate(csv_files):
        df = pd.read_csv(item, sep=';',  dtype={'Quantity': float})
        final_part_name = ' '.join(os.path.splitext(os.path.splitext(item)[0])[0].split('\\')[1].split(' ')[0:3])
        df.to_excel(os.path.join(transient_xlsx_file, "{}_{}.xlsx".format(final_part_name, index)), engine='openpyxl',
                    index=False)


def report_diff(x):
    """Function to highlight value change"""
    # return x[0] if x[0] == x[1] else '{} ---> {}'.format(*x)
    if x[0] == x[1]:
        return x[0]
    elif pd.isnull(x[0]):
        return x[0]
    elif x[0] > x[1]:
        return '{} ---> {}?'.format(*x)
    else:
        return '{} ---> {}!'.format(*x)


def comparic(file1, file2):
    """
    Function which process xlsx files from indicated directory, compare them and finally
    presents changes between versions.

    Args:
        file1 (str): first xlsx file path
        file2 (str): second xlsx file path
    """

    # Setting global variables which would be returned from function in the end
    global df_changed, removed, added

    # Loading dfs from excels files located in pointed directory
    previous = pd.read_excel(file1, engine="openpyxl", na_values=['NA']).dropna(how='all')
    present = pd.read_excel(file2, engine="openpyxl",  na_values=['NA']).dropna(how='all')

    # Need to delete remaining empty rows in dfs
    previous = previous.dropna(subset=["Component Description"])
    present = present.dropna(subset=["Component Description"])

    # Setting index column/columns
    index_col = ["Component Description", 'Nominal Size']
    previous_df = previous.set_index(index_col)
    present_df = present.set_index(index_col)

    # First divide the df for those elements which were removed or added from/to present df
    previous_keys = previous_df.index
    present_keys = present_df.index

    if isinstance(previous_keys, pd.MultiIndex):
        removed_keys = previous_keys.difference(present_keys)
        added_keys = present_keys.difference(previous_keys)
    else:
        removed_keys = np.setdiff1d(previous_keys, present_keys)
        added_keys = np.setdiff1d(present_keys, previous_keys)

    out_data = {}

    removed = previous_df.loc[removed_keys]
    if not removed.empty:
        out_data["removed"] = removed

    added = present_df.loc[added_keys]
    if not added.empty:
        out_data["added"] = added

    # return keys and columns which are common for previous and present df
    common_keys = np.intersect1d(previous_keys, present_keys, assume_unique=True)
    common_columns = np.intersect1d(previous_df.columns, present_df.columns, assume_unique=True)

    new_common = present_df.loc[common_keys, common_columns]
    old_common = previous_df.loc[common_keys, common_columns]

    # Concat common data from both dfs and dropping duplicates
    common_data = pd.concat([old_common.reset_index(), new_common.reset_index()], sort=True)
    changed_keys = common_data.drop_duplicates(keep=False)[index_col]

    if isinstance(changed_keys, pd.Series):
        changed_keys = changed_keys.unique()
    else:
        changed_keys = changed_keys.drop_duplicates().set_index(index_col).index

    # Concat of this items which were common in both dfs, swaping columns and then first format
    df_all_changes = pd.concat([old_common.loc[changed_keys], new_common.loc[changed_keys]], axis='columns',
                               keys=['old', 'new']).swaplevel(axis=1)

    df_changed = df_all_changes.groupby(level=0, axis=1).apply(lambda frame: frame.apply(report_diff, axis=1))

    # creating a df with changed items
    if not df_changed.empty:
        out_data['changed'] = df_changed

    # reset index to attain as similar as original df item schedule
    df_changed = df_changed.reset_index()
    added = added.reset_index()
    removed = removed.reset_index()

    # set sorting for dfs
    df_changed = df_changed.sort_values(by=['Sheet name'])
    added = added.sort_values(by=["Sheet name"])
    removed = removed.sort_values(by=["Sheet name"])

    return df_changed, added, removed


def semi_final_excel(file_1, file_2, semi_final_output):
    """Function to export output from comparic() function to excel"""
    global writer
    df3 = comparic(file_1, file_2)
    output_columns = ["Nominal Size", "Component Description", "Sch. ", "Rating", "Spec", "Quantity", "Note",
                      "Sheet name"]
    writer = pd.ExcelWriter(semi_final_output, engine='xlsxwriter')
    df_changed.to_excel(writer, "changed", index=False, columns=output_columns)
    added.to_excel(writer, "removed", index=False, columns=output_columns)
    removed.to_excel(writer, "added", index=False, columns=output_columns)
    writer.save()
    return writer


def colour_my_excel(semi_final, format_semi_final):
    """Format pointed excel file by usage of 'openpyxl' library"""
    workbook = load_workbook(filename=semi_final)

    # SHEETNAMES
    changed = workbook['changed']
    removed = workbook['removed']
    added = workbook['added']

    # COLOURS
    orange_colour = 'ffa500'
    red_colour = 'ffc7ce'
    yellow_colour = 'ffff00'
    green_colour = '1ea51e'

    # PATTERNS
    yellow_fill = styles.PatternFill(start_color=yellow_colour, end_color=yellow_colour, fill_type='solid')
    orange_fill = styles.PatternFill(start_color=orange_colour, end_color=orange_colour, fill_type='solid')
    red_background = PatternFill(start_color=red_colour, end_color=red_colour, fill_type='solid')
    green_background = PatternFill(start_color=green_colour, end_color=green_colour, fill_type='solid')

    # RULES
    rule_yellow = Rule(type='containsText', text='?', stopIfTrue=True)
    rule_orange = Rule(type='containsText', text='!', stopIfTrue=True)

    # RULE.DXF
    rule_yellow.dxf = DifferentialStyle(border=None, fill=yellow_fill)
    rule_orange.dxf = DifferentialStyle(border=None, fill=orange_fill)

    # CONDITIONAL FORMATTING
    changed.conditional_formatting.add('A1:H150', rule_yellow)
    changed.conditional_formatting.add('A1:H150', rule_orange)

    # STYLE FOR REMOVED ROWS
    row_range = removed.max_row
    for rows in removed.iter_rows(min_row=2, max_row=row_range, min_col=1):
        for cell in rows:
            cell.fill = red_background

    # STYLE FOR ADDED ROWS
    row_range = added.max_row
    for rows in added.iter_rows(min_row=2, max_row=row_range, min_col=1):
        for cell in rows:
            cell.fill = green_background

    workbook.save(format_semi_final)


def run_the_script():
    input_data = "input_data"
    transient_csv_directory = "transient_csv_directory"
    transient_xlsx_directory = "transient_xlsx_directory"
    semi_final_excel_name = "output_data/semi_final.xlsx"
    format_semi_final = "output_data/semi_final_formated.xlsx"
    both_data_csv_output(input_data, transient_csv_directory)
    csv_to_excel(transient_csv_directory, transient_xlsx_directory)
    semi_final_excel(os.path.join(transient_xlsx_directory, 'Material Take Off_0.xlsx'),
                     os.path.join(transient_xlsx_directory, 'Material Take Off_1.xlsx'), semi_final_excel_name)
    colour_my_excel(semi_final_excel_name, format_semi_final)


run_the_script()
